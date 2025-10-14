from fpdf import FPDF
from openpyxl import Workbook

def generar_pdf(empleado, contratos):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Reporte de contratos de {empleado}", ln=True, align='C')
    pdf.ln(10)
    for c in contratos:
        pdf.cell(200, 10, txt=f"{c['fecha_inicio']} - {c['fecha_fin']} | ${c['valor_contrato']}", ln=True)
    pdf.output(f"reporte_{empleado}.pdf")
    return f"reporte_{empleado}.pdf"


def generar_excel(empleado, contratos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    ws.append(["Empleado", "Fecha inicio", "Fecha fin", "Valor contrato"])
    for c in contratos:
        ws.append([empleado, c['fecha_inicio'], c['fecha_fin'], c['valor_contrato']])
    file_name = f"reporte_{empleado}.xlsx"
    wb.save(file_name)
    return file_name
